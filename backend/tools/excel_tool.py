from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import re


def generate_xlsx(content: str, style_params: dict, file_path: str) -> str:
    """Generate an Excel spreadsheet from content with style parameters."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    font_family = style_params.get("font_family", "Calibri")
    font_size = style_params.get("font_size", 12)
    text_color = style_params.get("text_color", "000000")

    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        data = _parse_table_from_text(content)

    _populate_worksheet(ws, data, font_family, font_size, text_color)

    wb.save(file_path)
    return file_path


def _parse_table_from_text(text: str) -> list:
    """Parse table-like structure from text content."""
    lines = text.strip().split("\n")
    data = []

    for line in lines:
        if not line.strip():
            continue

        if "|" in line:
            row = [cell.strip() for cell in line.split("|") if cell.strip()]
            data.append(row)
        elif "," in line:
            row = [cell.strip() for cell in line.split(",")]
            data.append(row)
        else:
            data.append([line.strip()])

    return data if data else [["Content"], [text]]


def _populate_worksheet(ws, data: list, font_family: str, font_size: int, text_color: str):
    """Populate worksheet with data and styling."""
    rgb = _hex_to_rgb(text_color)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name=font_family, size=font_size, bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    data_font = Font(name=font_family, size=font_size, color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}")

    for row_idx, row_data in enumerate(data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = data_font

    for col_idx in range(1, len(data[0]) + 1 if data else 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def _hex_to_rgb(hex_color: str):
    """Convert hex color to RGB tuple."""
    hex_color = hex_color.lstrip("#")
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
